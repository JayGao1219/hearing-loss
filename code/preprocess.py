import openpyxl
import os

def process_excel(input_path, output_path):
    # 用于处理文件 audiogram.xlsx
    # 这个文件有*
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(input_path)

    # 新建一个 Excel 文件
    output_workbook = openpyxl.Workbook()
    output_worksheet = output_workbook.active
    iid=[]

    # 遍历每一行
    for worksheet in workbook:
        print(worksheet.title)
        for row in worksheet.iter_rows(values_only=True):
            row = list(row)
            if row[0] in iid and type(row[0]) is not str:
                row[0] = -row[0]
            # 如果只有最后一列是 **，则赋值为倒数第二列的值
            if row[-1] == '**' and row[-2] != '**':
                row[-1] = row[-2]
            # 如果非边界值是 **，则进行插值
            if '**' in row[1:-1]:
                continue
            if isinstance(row[0],str):
                continue

            # 写入新的 Excel 文件
            iid.append(row[0])
            row = tuple(row)
            output_worksheet.append(row)

    # 保存并关闭文件
    output_workbook.save(output_path)


if __name__=='__main__':
    process_excel("../data/raw_data/audiogram.xlsx","../data/audiogram.xlsx")